import os
import re
import argparse
import numpy as np
import pandas as pd
from shutil import copy2
from afcpy.analysis import Dataset
from afcpy.models import Model1, Model2
from openpyxl import load_workbook, Workbook

def fetch():
    """
    pulls new data files from the network drive
    """
    
    # hard-coded directories.
    repo_network = 'Z:/Behavior/Data/Andrew/'
    # repo_local = 'C:/Users/jbhunt/Dropbox/Lab/Felsen/Data/Josh/'
    repo_local = '/home/jbhunt/Dropbox/Lab/Felsen/Data/Josh/'
    # repo_network = 'C:/Users/jbhunt/Desktop/net/'
    # repo_local = 'C:/Users/jbhunt/Desktop/loc/'
    
    # experiment labels
    labels = ('glu',
              'arch',
              'gadCTL',
              'gluCTL')
    
    # collect all data files in the network repo.
    data_net = []
    
    for root, folders, files in os.walk(repo_network):
        for file in files:
            if (file.startswith('data'))&(any(label in file for label in labels)):
                
                animal_id = file.split('Andrew_')[-1].split('_')[0] # get the animal ID
                date = re.findall('\d{6}\w{1}',file)[0] # get the date
                ffp = os.path.join(root,file)
                
                data_net.append([animal_id,date,ffp])
    
    # collect all data files in the local repo
    data_loc = []
    
    for root, folders, files in os.walk(repo_local):
        for file in files:
            if file.startswith('data'):
                
                animal_id = file.split('Andrew_')[-1].split('_')[0] # get the animal ID
                date = re.findall('\d{6}\w{1}',file)[0] # get the date
                ffp = os.path.join(root,file)
                 
                data_loc.append([animal_id,date])
    
    data_new = []
    print('\nNew files:\n-------------------')
    
    for entry in data_net:
        data = entry[:2]
        if data not in data_loc:
            animal_id = entry[0]
            ffp = entry[-1]
            data_new.append((animal_id,ffp))
            basename = os.path.basename(ffp)
            print('{}'.format(basename))
    
    print('\n')
            
    proceed = input('These files will be transferred to the local repository. Would you like to proceed? ')
    if (proceed == 'y')|(proceed == 'yes'):
        pass
    elif (proceed == 'n')|(proceed == 'no'):
        return
    else:
        print('{} is not a valid option. Exiting program.'.format(proceed))
        return
    
    n_files = len(data_new)
    bars_per_file = int(100/n_files)*'|'
    for i,(animal_id,ffp) in enumerate(data_new):
        dst = os.path.join(repo_local,animal_id)
        copy2(ffp,dst)
        
    print('\nDone!')
    
def process(**kwargs):
    """
    generates taskbase files for unprocessed data files in the local repository
    """
    
    _kwargs = {'exclude': [190829,190830,190831], # dates to be excluded
               }
    
    # repo_loc = 'C:/Users/jbhunt/Dropbox/Lab/Felsen/Data/Josh/'
    repo_loc = '/home/jbhunt/Dropbox/Lab/Felsen/Data/Josh/'
    
    # collect all data files in the local repo
    data_loc = []
    tb_loc = []
    
    for root, folders, files in os.walk(repo_loc):
        for file in files:
            if file.startswith('data'):
                animal_id = file.split('Andrew_')[-1].split('_')[0] # get the animal ID
                date = re.findall('\d{6}\w{1}',file)[0] # get the date
                ffp = os.path.join(root,file)
                if int(date[:-1]) in _kwargs['exclude']:
                    continue
                else:
                    data_loc.append([animal_id,date,ffp])
                
            if file.startswith('tb'):
                animal_id = file.split('Andrew_')[-1].split('_')[0] # get the animal ID
                date = re.findall('\d{6}\w{1}',file)[0] # get the date
                ffp = os.path.join(root,file)
                tb_loc.append([animal_id,date])
    
    data_new = []
    
    for entry in data_loc:
        info = entry[:2]
        if info not in tb_loc:
            animal_id = entry[0]
            ffp = entry[-1]
            data_new.append((animal_id,ffp))
    
    print('\nNew data files:\n----------')    
    for entry in data_new:
        basename = os.path.basename(entry[-1])
        print(basename)
        
    proceed = input('\nThese files will be processed. Would you like to proceed? ')
    if (proceed == 'y')|(proceed == 'yes'):
        pass
    elif (proceed == 'n')|(proceed == 'no'):
        return
    else:
        print('\n{} is not a valid response. Exiting program.'.format(proceed))
        return
            
    eng = engine.start_matlab()
    
    for entry in data_new:
        src = entry[-1]
        animal_id = entry[0]
        dir = os.path.join(repo_loc,animal_id)
        basename = os.path.basename(src)
        file_info = basename.split('data_')[-1]
        tb_file = 'tb_'+file_info
        dst = os.path.join(dir,tb_file)
        eng.generate_taskbase(src,dst)

def push():
    """
    pushes newly generated taskbase files to the github repository
    """
    
    # repo_loc = 'C:/Users/jbhunt/Dropbox/Lab/Felsen/Data/Josh'
    repo_loc = '/home/jbhunt/Dropbox/Lab/Felsen/Data/Jacki/'
    # repo_git = 'C:/Users/jbhunt/Dropbox/Code/Python/lib/afcpy/afcpy/data/josh/'
    repo_git = '/home/jbhunt/Dropbox/Code/Python/lib/afcpy/afcpy/data/jacki/'
    
    tb_files = []
    
    for root,folders,files in os.walk(repo_loc):
        for file in files:
            if file.startswith('tb'):
                tb_files.append(os.path.join(root,file))
    
    print('\nNew taskbase files:\n----------')
    for tb_file in tb_files:
        basename = os.path.basename(tb_file)
        dirname = os.path.dirname(tb_file)
        animal_id = os.path.basename(dirname)
        animal_dir = os.path.join(repo_git,animal_id)
        ffp = os.path.join(repo_git,animal_id,basename)
        
        if os.path.exists(ffp):
            continue
        
        else:
            print('transferring {} to github repo ...'.format(basename))
            src = tb_file
            dst = animal_dir
            copy2(src,dst)
            
    print('\nDone!')
    
### processing Jacki's modeling results ###

def process_model_datasets(src, tag=None):
    """
    """
    
    excel_files = []
    for root,folders,files in os.walk(src):
        for file in files:
            ffp = os.path.join(root,file)
            if tag is None:
                excel_files.append(ffp)
            elif tag in file:
                excel_files.append(ffp)
            else:
                continue
                
    for xlsx in excel_files:
        print(xlsx)
        ds = ModelDataset(xlsx)
        ds._update_excel_file()
        
def combine_spreadsheets(src, dst):
    """
    """
    
    for root,folders,files in os.walk(src):
        if len(folders) > 0:
            print(folders)
            break
        
    # make sure the first folder in the list corresponds to sessions 1-20
    folders = sorted(folders)
        
    proj_name = os.path.basename(root)
    dir1,dir2 = [os.path.join(root,folder) for folder in folders]
    
    files1 = []
    for root,folders,files in os.walk(dir1):
        for file in files:
            ffp = os.path.join(root,file)
            files1.append(ffp)
    
    files2 = []
    for root,folders,files in os.walk(dir2):
        for file in files:
            ffp = os.path.join(root,file)
            files2.append(ffp)
            
    pairs =  [pair for pair in zip(sorted(files1),sorted(files2))]
    
    for pair in pairs:
        wb0 = Workbook() # the new workbook
        wb1 = load_workbook(pair[0])
        wb2 = load_workbook(pair[1])
        wb_number = ''.join([i for i in os.path.basename(pair[0]) if i.isdigit()])
        new_file_base_name = proj_name + wb_number + '.xlsx'
        
        for sheetname in wb1.sheetnames:
            if sheetname == 'analysis':
                continue
            sheet = wb1[sheetname]
            wb0.create_sheet(sheetname)
            for row in sheet:
                for cell in row:
                    wb0[sheetname][cell.coordinate].value = cell.value
                    
        for sheetname in wb2.sheetnames:
            if sheetname == 'analysis':
                continue
            sheet = wb2[sheetname]
            sheet_number = int(''.join([i for i in sheetname if i.isdigit()]))
            new_sheetname = 'Sheet{}'.format(sheet_number + 20)
            wb0.create_sheet(new_sheetname)
            for row in sheet:
                for cell in row:
                    wb0[new_sheetname][cell.coordinate].value = cell.value
                    
        dstdir = os.path.join('/home/jbhunt/Desktop',dst)
        wb0.save(os.path.join(dstdir,new_file_base_name))
            
if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--sync',help='fetch new files',action='store_true')
    parser.add_argument('--analyze',help='analyze model results',action='store_true')
    parser.add_argument('--src',help='dir which contains the model results',default='',type=str)
    parser.add_argument('--tag',help='tag which identifies csv files to analyze',default='GABA',type=str)
    parser.set_defaults(sync=False,analyze=False)
    args = parser.parse_args()
    
    if args.sync:
        import matlab.engine as engine # this import here because it's not installed on Ubuntu
        fetch()
        process()
        push()
        
    if args.analyze:
        process_model_datasets(args.src,args.tag)