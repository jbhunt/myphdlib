import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from afcpy.analysis import Dataset

def output_animal_results(dst, test_h0=False, include_oe=True):
    """
    analyzes all datasets and writes the results to a csv file

    keywords
    --------
    dst : str
        the filepath/name for the csv file
    test_h0 : bool (default is False)
        flag for performing the boostrap hypothesis testing
    include_oe : bool (default is True)
        flag for including out-early trial-types
    """

    combos = (['VGlut2-cre','ChR2','blue'],
              ['VGlut2-cre','none','blue'],
              ['Gad2-cre','ChR2','blue'],
              ['Gad2-cre','Arch','green'],
              ['Gad2-cre','none','blue'],
              ['Gad2-cre','none','green'],
              ['wildtype','none','none'],
              )

    columns = ('filename',
               'bias_curve_fit',
               'bias_beta_coef',
               'boot_result_p',
               'mwu_result_left_u',
               'mwu_result_left_p',
               'mwu_result_right_u',
               'mwu_result_right_p',
               'med_rxn_time_manip_off_left',
               'med_rxn_time_manip_on_left',
               'med_rxn_time_manip_off_right',
               'med_rxn_time_manip_on_right',
               )

    df = pd.DataFrame(dict(),columns=columns)

    for combo in combos:

        # init the dataset
        ds = Dataset()
        ds.load(*combo,include_oe=include_oe)

        # perform bootstrap hypothesis testing
        if test_h0:
            ds.test(n_iters=5000)
        else:
            ds.p_values = np.full(len(ds.fileset),np.nan)

        # if analyzing the muscimol experiment extract the triplet of filenames
        if combo == ('wildtype','none','none'):
            basenames = []
            for trip in ds.fileset:
                trip_basenames = [os.path.basename(f) for f in trip]
                basenames.append(trip_basenames)
        else:
            basenames = [os.path.basename(f) for f in ds.fileset]

        df_data = (basenames,
                   ds.b1,
                   ds.b2,
                   ds.p_values,
                   ds.rt_effect_left_u,
                   ds.rt_effect_left_p,
                   ds.rt_effect_right_u,
                   ds.rt_effect_right_p,
                   ds.rt_iv_0_left,
                   ds.rt_iv_1_left,
                   ds.rt_iv_0_right,
                   ds.rt_iv_1_right
                   )
        df_temp = pd.DataFrame(np.array(df_data).T, columns=columns)
        df = df.append(df_temp)

    df.to_csv(dst)

def get_reaction_time_data(dataset, filename=None, x_manip='x_light'):
    """
    """

    fontStyle = Font(name="Calibri", size=12, color=colors.BLACK)

    if type(dataset) == list:
        sessions = dataset
        filenames = [os.path.basename(s.session_id).rstrip('.mat') for s in sessions]
        idx = np.array([[x_manip,'x_choice'],[0,0],[0,1],[1,0],[1,1]])

    elif type(dataset) == Dataset:

        sessions = dataset.sessions

        if type(sessions[0].session_id) == list:
            # filenames = ['trip{}'.format(itrip) for itrip in range(len(dataset.sessions))]
            filenames = [os.path.basename(s.session_id[0]) for s in dataset.sessions]
            idx = np.array([['x_other','x_choice'],[0,0],[0,1],[1,0],[1,1]])

        else:
            filenames = [os.path.basename(s.session_id).rstrip('.mat') for s in sessions]
            idx = np.array([[x_manip,'x_choice'],[0,0],[0,1],[1,0],[1,1]])

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    for isession,s in enumerate(sessions):
        label = filenames[isession]
        try:
            sheet = wb.create_sheet('session{}'.format(isession))
        except:
            import pdb; pdb.set_trace()
        sheet.cell(1,1).value = label
        sheet.cell(1,1).font = fontStyle
        for (i,j),z in np.ndenumerate(np.zeros([5,2])):
            cell = sheet.cell(i + 2,j + 1)
            cell.value = idx[i,j]
            cell.font = fontStyle

        row1 = s.data[(s.data[x_manip]==0) & (s.data.x_choice==0)]['x_reaction_time']
        row2 = s.data[(s.data[x_manip]==0) & (s.data.x_choice==1)]['x_reaction_time']
        row3 = s.data[(s.data[x_manip]==1) & (s.data.x_choice==0)]['x_reaction_time']
        row4 = s.data[(s.data[x_manip]==1) & (s.data.x_choice==1)]['x_reaction_time']

        for irow,row in enumerate([row1,row2,row3,row4]):
            for icol in range(len(row)):
                cell = sheet.cell(irow + 3,icol + 3)
                cell.value = row.iloc[icol]
                cell.font = fontStyle

    if filename is None:
        filename = input('Please enter a filename for the workbook: ')

    wb.save(filename)
    wb.close()
    print('All done!')

    return filename
